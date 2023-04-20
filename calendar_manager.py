from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import datetime
import peopleManagement

def getNextMonthPath():
    path = "excel/"
    month = datetime.date.today().month
    nextMonth = getNextMonth(month)
    date = datetime.date.today()
    path = path + getMonthString(nextMonth) + str(date.year) + ".xlsx"
    return path

def getThisMonthPath():
    path = "excel/"
    month = datetime.date.today().month
    date = datetime.date.today()
    path = path + getMonthString(month) + str(date.year) + ".xlsx"
    return path

def getMonthString(number):
    if(number == 1):
        return "January"
    if(number == 2):
        return "February"
    if(number == 3):
        return "March"
    if(number == 4):
        return "April"
    if(number == 5):
        return "May"
    if(number == 6):
        return "June"
    if(number == 7):
        return "July"
    if(number == 8):
        return "August"
    if(number == 9):
        return "September"
    if(number == 10):
        return "October"
    if(number == 11):
        return "November"
    if(number == 12):
        return "December"

def getNextMonth(month):
    if(month == 12):
        return 1
    else:
        return month + 1

path = getNextMonthPath()

def getCalendarDays(path):
    wb = load_workbook(path)
    ws = wb.active
    days = []
    for col in range(2, 35):
        day = {}
        for row in range(3, 10):
            char = get_column_letter(col)
            name = ws['A' + str(row)].value
            if(name != None):
                weekday = ws[char + '2'].value
                date = ws[char + '1'].value
                if(weekday != None):
                    value = ws[char + str(row)].value
                    if(value!= None):
                        day[name] = {
                            'name': name,
                            'weekday': weekday,
                            'date': date,
                            'value': value
                        }
                    else:
                        day[name] = {
                            'name': name,
                            'weekday': weekday,
                            'date': date,
                            'value': ''
                        }
        if(day != {}):
             days.append(day)
    return days

days = getCalendarDays(path)
#
# for day in days:
#     if 'man' in day:
#         print(day['man'])


