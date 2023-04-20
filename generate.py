# it generates the empty excel of the next month

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import datetime
import peopleManagement

def getMonthString(number):
    if(number == 1):
        return "January"
    if(number == 2):
        return "February"
    if(number == 3):
        return "March"
    if(number == 4):
        return "April"
    if(number == 5):
        return "May"
    if(number == 6):
        return "June"
    if(number == 7):
        return "July"
    if(number == 8):
        return "August"
    if(number == 9):
        return "September"
    if(number == 10):
        return "October"
    if(number == 11):
        return "November"
    if(number == 12):
        return "December"

def getWeekDay(day):
    if(day == 0 or day == 7):
        return "Monday"
    if(day == 1):
        return "Tuesday"
    if(day == 2):
        return "Wednesday"
    if(day == 3):
        return "Thursday"
    if(day == 4):
        return "Friday"
    if(day == 5):
        return "Saturday"
    if(day == 6):
        return "Sunday"

def getNextMonth(month):
    if(month == 12):
        return 1
    else:
        return month + 1

month = datetime.date.today().month

# print(datetime.date.today() + datetime.timedelta(days=30))
path = "excel/"

wb = Workbook()
ws = wb.active

myMonth = datetime.date.today().month
nextMonth = getNextMonth(myMonth)

daysAfter = 0
run = True
col = 1
ws["A2"] = "Name \ Date"
ws["A2"].font = Font(bold=True)
while run:
    daysAfter = daysAfter + 1
    date = datetime.date.today() + datetime.timedelta(days=daysAfter)
    if(date.month != myMonth and date.month != nextMonth):
        run = False
        break
    if(date.month == nextMonth):
        col = col + 1
        char = get_column_letter(col)
        ws[char + "1"] = date
        ws[char + "1"].font = Font(bold=True, italic=True, size=9)
        ws[char + "2"] = getWeekDay(date.weekday())
        ws[char + "2"].font = Font(bold=True)

staffPath = "excel/people.xlsx"
myStaff = peopleManagement.Staff(staffPath)
people = myStaff.people

for person in people:
    ws.append([person["Name"]])

wb.save(path + getMonthString(nextMonth) + str(date.year) + ".xlsx")
