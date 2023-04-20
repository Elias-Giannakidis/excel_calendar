from Person import Person, genPeople
from peopleManagement import Staff, getPossibelDays
from openpyxl import Workbook, load_workbook
from calendar_manager import getCalendarDays, getNextMonthPath, getThisMonthPath
from openpyxl.utils import get_column_letter
import itertools
import random

# Get the days of the calendar of the next month
monthPath = getNextMonthPath()
days = getCalendarDays(monthPath)

# Count how many days the month has
monthDays = 0
for day in days:
    monthDays = monthDays + 1

staff_path = "excel/people.xlsx"
pdays_path = 'excel/vardies.xlsx'
myStaff = Staff(path=staff_path)

# People of the shop
people = genPeople(myStaff.people)

# Define the person calendar
for person in people:
    calendar = []
    for day in days:
        if person.name in day:
            calendar.append(day[person.name])
    person.calendar = calendar

# print people!
# for person in people:
#     person.printMe()

# An array with arrays with the possible work days - vardies
vardies = getPossibelDays(pdays_path)
newVardies = []
for vardia in vardies:
    posVardies = itertools.permutations(vardia)
    for posVardia in posVardies:
     newVardies.append(posVardia)
vardies = newVardies

# Count the possible vardies
vardiesNumber = 0
for _ in vardies:
    vardiesNumber = vardiesNumber + 1

# getCalendar score
def getCalendarScore(people, calendar):
    j = 0
    score = 0
    for person in people:
        newCalendar = []
        for day in calendar:
            newCalendar.append(day[j])
        j = j + 1
        score = score + person.getScore(newCalendar)
    return score


#
# test = sorted(people, key=lambda person: -1 * person.getScore())
# test = test[:50]
#

def getRandomVardiaArray():
    vardiaArray = []
    for _ in range(monthDays):
        vardiaArray.append(random.randint(0, vardiesNumber - 1))
    return vardiaArray

def makeCalendarFromVardiaArray(array):
    calendar = []
    for i in array:
        calendar.append(vardies[i])
    return calendar

def neurosMethod():
    # init calendars
    a = 1000
    # children
    c = 1000
    # Survive
    s = 10
    # possibility to change a row at 100
    p = 30
    # Epochs
    e = 100
    calendars = []
    for _ in range(a):
        vardiaArray = getRandomVardiaArray()
        calendar = makeCalendarFromVardiaArray(vardiaArray)
        # score = getCalendarScore(people, calendar)
        calendars.append(calendar)
    for __ in range(e):
        for calendar in calendars:
            newCalendars = calendars.copy()
            for _ in range(c):
                newCalendar = calendar.copy()
                for i in range(monthDays):
                    if(random.randint(0, 100) < p):
                        newCalendar[i] = vardies[random.randint(0, vardiesNumber - 1)]
                newCalendars.append(newCalendar)
        calendars = newCalendars
        calendars = sorted(calendars, key=lambda calendar: -1 * getCalendarScore(people, calendar))
        calendars = calendars[:s]
        print('best score: ', getCalendarScore(people, calendars[0]))
    return calendars



# Loop to make the possible calendars
calendars = [[]]
count = 0
for _ in range(monthDays):
    count = count + 1
    newCalendars = []
    for vardia in vardies:
        for day in calendars:
            newDay = day.copy()
            newDay.append(vardia)
            newCalendars.append(newDay)
    newCalendars = sorted(newCalendars, key=lambda calendar: -1 * getCalendarScore(people, calendar))
    if(count % 3 == 0):
        newCalendars = newCalendars[:2]
    calendars = newCalendars
# for i in range(10):
#     print(calendars[i])
calendars = neurosMethod()
program1 = calendars[1]
# i = 0
# for day in program1:
#     print(days[i]['man']['weekday'], "::", day)
#     i = i + 1

def makeExcel(index):
    path = monthPath
    program = calendars[index]
    wb = load_workbook(path)
    ws = wb.active
    # make the first column
    for col in range(2, monthDays + 2):
        char = get_column_letter(col)
        dayProgramm = program[col - 2]
        i = 3
        for personDayProgram in dayProgramm:
            ws[char + str(i)] = personDayProgram
            i = i + 1
    wb.save(path)

makeExcel(0)

# for calendar in calendars:
#     print(calendar)

