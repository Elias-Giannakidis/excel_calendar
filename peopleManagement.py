from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

path = "excel/people.xlsx"

class Staff:
    def __init__(self, path):
        self.path = path
        self.wb = load_workbook(path)
        self.ws = self.wb.active
        self.countPeople()
        self.getConsts()
        self.getPersons()

    def countPeople(self):
        self.names = []
        self.peopleNum = 0
        for i in range(2, 10):
            name = self.ws['A' + str(i)].value
            if (name == None):
                break
            self.peopleNum = i - 1
            self.names.append(name)

    def getConsts(self):
        self.consts = []
        for col in range(1, 30):
            char = get_column_letter(col)
            const = self.ws[char + '1'].value
            if(const == None):
                break
            self.consts.append(const)


    def getPersons(self):
        self.people = []
        for i in range(2, self.peopleNum + 2):
            col = 1
            person = {}
            for const in self.consts:
                char = get_column_letter(col)
                value = self.ws[char + str(i)].value
                col = col + 1
                if(value != None):
                    person[const] = value
            self.people.append(person)

def getPossibelDays(path):
    wb = load_workbook(path)
    ws = wb.active
    days = []
    col = 0
    while True:
        col = col + 1
        char = get_column_letter(col)
        run = True
        day = []
        row = 0
        while run:
            row = row + 1
            vardia = ws[char + str(row)].value
            if(vardia == None):
                run = False
            else:
                day.append(vardia)
        if(day == []):
            break
        days.append(day)
    return days


